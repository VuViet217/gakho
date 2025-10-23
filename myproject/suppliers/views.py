from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from .models import Supplier, PurchaseOrder, PurchaseOrderItem
from .forms import SupplierForm, PurchaseOrderForm, PurchaseOrderItemForm
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import io

# Supplier Views
@login_required
def supplier_list(request):
    search_query = request.GET.get('search', '')
    suppliers = Supplier.objects.all()
    
    if search_query:
        suppliers = suppliers.filter(
            Q(code__icontains=search_query) | 
            Q(name__icontains=search_query) | 
            Q(address__icontains=search_query) |
            Q(contact_name__icontains=search_query) |
            Q(contact_phone__icontains=search_query) |
            Q(contact_email__icontains=search_query)
        )
    
    paginator = Paginator(suppliers, 10)  # 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'title': 'Danh sách nhà cung cấp'
    }
    return render(request, 'suppliers/supplier_list.html', context)


@login_required
def supplier_create(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nhà cung cấp đã được tạo thành công!')
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    
    return render(request, 'suppliers/supplier_form.html', {
        'form': form,
        'title': 'Thêm nhà cung cấp mới'
    })


@login_required
def supplier_update(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nhà cung cấp đã được cập nhật thành công!')
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    
    return render(request, 'suppliers/supplier_form.html', {
        'form': form,
        'supplier': supplier,
        'title': 'Cập nhật nhà cung cấp'
    })


@login_required
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    
    if request.method == 'POST':
        try:
            supplier.delete()
            messages.success(request, 'Nhà cung cấp đã được xóa thành công!')
        except Exception as e:
            messages.error(request, f'Không thể xóa nhà cung cấp này. Lỗi: {str(e)}')
        return redirect('supplier_list')
    
    return render(request, 'suppliers/supplier_confirm_delete.html', {
        'supplier': supplier,
        'title': 'Xóa nhà cung cấp'
    })


@login_required
def supplier_detail(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    purchase_orders = supplier.purchase_orders.all()
    
    paginator = Paginator(purchase_orders, 5)  # 5 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'suppliers/supplier_detail.html', {
        'supplier': supplier,
        'page_obj': page_obj,
        'title': f'Thông tin nhà cung cấp: {supplier.name}'
    })


# Purchase Order Views
@login_required
def po_list(request):
    search_query = request.GET.get('search', '')
    supplier_id = request.GET.get('supplier', '')
    status = request.GET.get('status', '')
    
    purchase_orders = PurchaseOrder.objects.all()
    
    # Tìm kiếm
    if search_query:
        purchase_orders = purchase_orders.filter(
            Q(po_number__icontains=search_query) | 
            Q(supplier__name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Lọc theo nhà cung cấp
    if supplier_id:
        purchase_orders = purchase_orders.filter(supplier_id=supplier_id)
    
    # Lọc theo trạng thái
    if status:
        purchase_orders = purchase_orders.filter(status=status)
    
    # Phân trang
    paginator = Paginator(purchase_orders, 10)  # 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Danh sách nhà cung cấp cho bộ lọc
    suppliers = Supplier.objects.all()
    
    context = {
        'page_obj': page_obj,
        'suppliers': suppliers,
        'search_query': search_query,
        'selected_supplier': supplier_id,
        'selected_status': status,
        'title': 'Danh sách đơn đặt hàng'
    }
    return render(request, 'suppliers/po_list.html', context)


@login_required
def po_create(request):
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đơn đặt hàng đã được tạo thành công!')
            return redirect('po_list')
    else:
        form = PurchaseOrderForm()
    
    return render(request, 'suppliers/po_form.html', {
        'form': form,
        'title': 'Tạo đơn đặt hàng mới'
    })


@login_required
def po_update(request, pk):
    purchase_order = get_object_or_404(PurchaseOrder, pk=pk)
    
    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST, request.FILES, instance=purchase_order)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đơn đặt hàng đã được cập nhật thành công!')
            return redirect('po_list')
    else:
        form = PurchaseOrderForm(instance=purchase_order)
    
    return render(request, 'suppliers/po_form.html', {
        'form': form,
        'purchase_order': purchase_order,
        'title': 'Cập nhật đơn đặt hàng'
    })


@login_required
def po_delete(request, pk):
    purchase_order = get_object_or_404(PurchaseOrder, pk=pk)
    
    if request.method == 'POST':
        try:
            purchase_order.delete()
            messages.success(request, 'Đơn đặt hàng đã được xóa thành công!')
        except Exception as e:
            messages.error(request, f'Không thể xóa đơn đặt hàng này. Lỗi: {str(e)}')
        return redirect('po_list')
    
    return render(request, 'suppliers/po_confirm_delete.html', {
        'purchase_order': purchase_order,
        'title': 'Xóa đơn đặt hàng'
    })


@login_required
def po_detail(request, pk):
    purchase_order = get_object_or_404(PurchaseOrder, pk=pk)
    items = purchase_order.items.all()
    
    if request.method == 'POST':
        form = PurchaseOrderItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.purchase_order = purchase_order
            item.save()
            messages.success(request, 'Sản phẩm đã được thêm vào đơn hàng!')
            return redirect('po_detail', pk=purchase_order.pk)
    else:
        form = PurchaseOrderItemForm()

    return render(request, 'suppliers/po_detail.html', {
        'purchase_order': purchase_order,
        'items': items,
        'form': form,
        'title': f'Chi tiết đơn đặt hàng: {purchase_order.po_number}'
    })


@login_required
@require_POST
def po_item_delete(request, pk):
    item = get_object_or_404(PurchaseOrderItem, pk=pk)
    po_id = item.purchase_order.id
    
    try:
        item.delete()
        messages.success(request, 'Sản phẩm đã được xóa khỏi đơn hàng!')
    except Exception as e:
        messages.error(request, f'Không thể xóa sản phẩm khỏi đơn hàng. Lỗi: {str(e)}')
    
    return redirect('po_detail', pk=po_id)


@login_required
def po_item_update(request, pk):
    item = get_object_or_404(PurchaseOrderItem, pk=pk)
    po_id = item.purchase_order.id
    
    if request.method == 'POST':
        form = PurchaseOrderItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Chi tiết đơn hàng đã được cập nhật!')
            return redirect('po_detail', pk=po_id)
    else:
        form = PurchaseOrderItemForm(instance=item)
    
    return render(request, 'suppliers/po_item_form.html', {
        'form': form,
        'item': item,
        'purchase_order': item.purchase_order,
        'title': 'Cập nhật chi tiết đơn hàng'
    })


@login_required
def supplier_export_excel(request):
    """Xuất danh sách nhà cung cấp ra file Excel"""
    
    # Tạo workbook mới
    wb = Workbook()
    ws = wb.active
    ws.title = "Nhà cung cấp"
    
    # Định dạng header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    headers = [
        'Mã NCC',
        'Tên nhà cung cấp',
        'Địa chỉ',
        'Tên liên hệ',
        'Điện thoại',
        'Email',
        'Website',
        'Mã số thuế',
        'Ghi chú',
        'Trạng thái'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # Mã NCC
    ws.column_dimensions['B'].width = 30  # Tên
    ws.column_dimensions['C'].width = 40  # Địa chỉ
    ws.column_dimensions['D'].width = 20  # Tên liên hệ
    ws.column_dimensions['E'].width = 15  # Điện thoại
    ws.column_dimensions['F'].width = 25  # Email
    ws.column_dimensions['G'].width = 20  # Website
    ws.column_dimensions['H'].width = 15  # MST
    ws.column_dimensions['I'].width = 30  # Ghi chú
    ws.column_dimensions['J'].width = 12  # Trạng thái
    
    # Lấy dữ liệu
    suppliers = Supplier.objects.all().order_by('code')
    
    # Định dạng cho data
    data_alignment = Alignment(vertical="center", wrap_text=True)
    
    # Ghi dữ liệu
    for row_num, supplier in enumerate(suppliers, 2):
        row_data = [
            supplier.code,
            supplier.name,
            supplier.address or '',
            supplier.contact_name or '',
            supplier.contact_phone or '',
            supplier.contact_email or '',
            supplier.website or '',
            supplier.tax_code or '',
            supplier.notes or '',
            'Hoạt động' if supplier.is_active else 'Ngừng'
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Tạo response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Danh_sach_nha_cung_cap_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def supplier_import_excel(request):
    """Import danh sách nhà cung cấp từ file Excel"""
    
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Vui lòng chọn file Excel để import!')
            return redirect('supplier_list')
        
        excel_file = request.FILES['excel_file']
        
        # Kiểm tra định dạng file
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'File phải có định dạng .xlsx hoặc .xls!')
            return redirect('supplier_list')
        
        try:
            # Đọc file Excel
            wb = load_workbook(excel_file)
            ws = wb.active
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Bỏ qua dòng header (row 1)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # Lấy dữ liệu từ các cột
                    code = row[0]
                    name = row[1]
                    address = row[2] if len(row) > 2 else None
                    contact_name = row[3] if len(row) > 3 else None
                    contact_phone = row[4] if len(row) > 4 else None
                    contact_email = row[5] if len(row) > 5 else None
                    website = row[6] if len(row) > 6 else None
                    tax_code = row[7] if len(row) > 7 else None
                    notes = row[8] if len(row) > 8 else None
                    is_active_text = row[9] if len(row) > 9 else 'Hoạt động'
                    
                    # Bỏ qua dòng trống
                    if not code or not name:
                        continue
                    
                    # Xử lý trạng thái
                    is_active = True
                    if is_active_text:
                        is_active = str(is_active_text).strip().lower() not in ['ngừng', 'ngung', 'inactive', 'false', '0']
                    
                    # Tạo hoặc cập nhật nhà cung cấp
                    supplier, created = Supplier.objects.update_or_create(
                        code=code,
                        defaults={
                            'name': name,
                            'address': address,
                            'contact_name': contact_name,
                            'contact_phone': contact_phone,
                            'contact_email': contact_email,
                            'website': website,
                            'tax_code': tax_code,
                            'notes': notes,
                            'is_active': is_active,
                        }
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Dòng {row_num}: {str(e)}')
            
            # Thông báo kết quả
            if success_count > 0:
                messages.success(request, f'Import thành công {success_count} nhà cung cấp!')
            
            if error_count > 0:
                error_msg = f'Có {error_count} lỗi khi import:<br>' + '<br>'.join(errors[:5])
                if len(errors) > 5:
                    error_msg += f'<br>... và {len(errors) - 5} lỗi khác'
                messages.warning(request, error_msg)
            
        except Exception as e:
            messages.error(request, f'Lỗi khi đọc file Excel: {str(e)}')
        
        return redirect('supplier_list')
    
    return redirect('supplier_list')


@login_required
def supplier_download_template(request):
    """Tải file Excel mẫu để import nhà cung cấp"""
    
    # Tạo workbook mới
    wb = Workbook()
    ws = wb.active
    ws.title = "Nhà cung cấp"
    
    # Định dạng header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    headers = [
        'Mã NCC*',
        'Tên nhà cung cấp*',
        'Địa chỉ',
        'Tên liên hệ',
        'Điện thoại',
        'Email',
        'Website',
        'Mã số thuế',
        'Ghi chú',
        'Trạng thái'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 12
    
    # Thêm dòng mẫu
    example_data = [
        'NCC001',
        'Công ty ABC',
        '123 Đường ABC, Quận 1, TP.HCM',
        'Nguyễn Văn A',
        '0901234567',
        'abc@company.com',
        'www.abc.com',
        '0123456789',
        'Ghi chú mẫu',
        'Hoạt động'
    ]
    
    data_alignment = Alignment(vertical="center", wrap_text=True)
    
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.alignment = data_alignment
        cell.border = thin_border
    
    # Thêm ghi chú
    note_row = 4
    ws.cell(row=note_row, column=1).value = "Ghi chú:"
    ws.cell(row=note_row, column=1).font = Font(bold=True)
    ws.cell(row=note_row+1, column=1).value = "* Cột bắt buộc"
    ws.cell(row=note_row+2, column=1).value = "Trạng thái: 'Hoạt động' hoặc 'Ngừng'"
    ws.cell(row=note_row+3, column=1).value = "Nếu mã NCC đã tồn tại, hệ thống sẽ cập nhật thông tin"
    
    # Tạo response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Mau_import_nha_cung_cap.xlsx"'
    
    wb.save(response)
    return response
